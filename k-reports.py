#!/usr/bin/env python3
import openpyxl
from datetime import datetime
import os
from settings import *

# 192.168.0.248', '-', 'servers', '[25/Mar/2019:01:18:25', '+0500]', '"CONNECT', 'https://195.122.177.135:443/', 'HTTP/1.0"', '200', '856
WEBTIMEOUT = 60  # время в минутах счётчика посещений
COUNT, SIZEB, TIME_STUMP, VISITS = (0, 1, 2, 3)
IP, USER, DATE, LINK, BYTES = (0, 2, 3, 6, 9)
COLUMN_WIDTH_REQUESTS = 10
COLUMN_WIDTH_VISITS = 10
# Стартовая колонка для записи данных
START_RECORD_COLUMN = 2
START_RECORD_COLUMN_LETTER = 'B'
# Стартовая строка для записи данных
START_RECORD_ROW = 3


def main():
    filename = choose_log()
    with open(filename, 'r') as fs:
        rows = fs.read().splitlines()

    users, dates = split_strings(rows)
    print("Пользователи:", users)
    print("Даты:", dates)
    try:
        print('Введите имя пользователя или \'{}\' если '
              'хотите выбрать всех пользователей...'.format(USERNAME_ALL_USERS))
        username = input("Введите Имя пользователя: ")
    except KeyboardInterrupt:
        exit()

    print('Текущий пользователь - ', username)
    if username not in users:
        print("Неверное имя пользователя")
        exit()

    statistics_dict, min_date, max_date = make_statistics(rows, username)
    generate_xls_report(REPORT_FILENAME, statistics_dict, min_date, max_date)
    return


def generate_xls_report(report_filename, statistics_dict, min_date, max_date):
    """
        Генерирует xlsx отчёт потребления интернет траффика для пользователя.
    """
    xlsx_workbook = openpyxl.Workbook()
    for user in sorted(statistics_dict.keys(), reverse=True):
        if user in EXCLUDE:
            continue
        data = statistics_dict[user]
        sheet = xls_head(xlsx_workbook, user, min_date, max_date)

        column_domain_length = 0
        column_traffic_size = 0
        xls_row = START_RECORD_ROW

        for domain_name, opts in sorted(data.items(), reverse=True, key=lambda x: x[1][1]):

            if opts[SIZEB] == 0:
                continue
            if len(domain_name) > column_domain_length:
                column_domain_length = len(domain_name)

            size_of_data = get_network_traffic_size(opts[SIZEB])
            if len(size_of_data) > column_traffic_size:
                column_traffic_size = len(size_of_data)

            xls_insert(sheet, xls_row, START_RECORD_COLUMN, domain_name, size_of_data, opts[COUNT], opts[VISITS])
            xls_row += 1

        xls_set_column_size(
            sheet,
            START_RECORD_COLUMN_LETTER,
            column_domain_length,
            column_traffic_size,
            COLUMN_WIDTH_REQUESTS,
            COLUMN_WIDTH_VISITS,
        )
    print("Отчёт от:", datetime.fromtimestamp(min_date), "до:", datetime.fromtimestamp(max_date))
    xlsx_workbook.save(report_filename)


def make_statistics(rows, username):
    """
        Создаёт словарь со статистикий по ссылкам, ключами которого являются имена пользователей, значением является
        словарь ключами которого являются доменные имена сайтов.
    """
    row = 0  # счётчик строк лога
    statistic_dict = {}  # словарь пользователя
    min_date = 999999999999999
    max_date = 0
    while row < len(rows):
        try:
            user = rows[row][USER]
            link = rows[row][LINK]
            link = parse_domain_name(link)
            byte = int(rows[row][BYTES])
            time_stump = time_to_timestump(rows[row][DATE])
            min_date = time_stump if time_stump < min_date else min_date
            max_date = time_stump if time_stump > max_date else max_date
            row += 1

            if username != "All":
                if user != username:
                    continue

            db_record = statistic_dict.get(user)
            if db_record:
                statistic_dict[user] = check_link_exists(db_record, link, byte, time_stump)
            else:
                # Make new user record - Count,Size,UnixTime,Visits
                links = {
                    link: [1, byte, time_stump, 1],
                }
                statistic_dict[user] = links
        except EOFError:
            break
        except ValueError as err:
            print(err)
            input('Press any key')
            continue

    return statistic_dict, min_date, max_date


def check_link_exists(db_record, link, byte, time_stump):
    """
        Функция ищет в словаре имя ссылки. Если находит, то добавляет - кол-во запросов, кол-во байт,
        кол-во посещений. Возвращает словарь с новыми значениями.
    """
    for db_link, options in db_record.items():
        if db_link == link:
            try:
                # Count links
                count = int(options[COUNT])
                count += 1
                # Size of link
                sizeb = int(options[SIZEB])
                sizeb += byte
                # Visits
                visits = int(options[VISITS])
                # TIME_STUMP
                db_timestump = int(options[TIME_STUMP])
                if time_stump - db_timestump > 60 * WEBTIMEOUT:
                    db_timestump = time_stump
                    visits += 1
                db_record[link] = [count, sizeb, db_timestump, visits]
            except ValueError as err:
                print(err)
                exit()
            return db_record
    else:
        db_record[link] = [1, byte, time_stump, 1]
    return db_record


def choose_log():
    """
        Ищет файлы срасширением *.log. Если находит, то просит указать индекс лога.
        Если не находит, то просит ввести имя файла.
        На выходе - имя файла для открытия
    """
    while True:
        files = [x for x in os.listdir(".") if x.endswith(".log")]
        if not files:
            filename = input("Введите имя файла с логом: ")
            if not filename:
                print('Введите корректное имя файла')
                continue
            return filename

        dict_of_log_files = {}
        file_index = 0
        for file_name in sorted(files, reverse=True):
            dict_of_log_files[file_index] = file_name
            print("{0}: {1}".format(file_index, file_name))
            file_index += 1

        try:
            index = int(input("Введите индекс лог-файла:"))
            filename = dict_of_log_files[index]
            return filename
        except (KeyError, ValueError):
            print("Введите корректный индекс лог-файла")
            continue


def xls_insert(sheet, xls_row, xls_start_col, *records):
    """
        sheet - страница, (xls_row, xls_start_col) - строка + стартовая колонка,
        Далее  *records содержит:
        имя ресурса, трафик с домена, кол-во запросов на ресурс, кол-во посещений за WEBTIMEOUT
    """
    xls_col = xls_start_col
    for record in records:
        cell = sheet.cell(row=xls_row, column=xls_col)
        cell.value = record
        xls_col += 1
    return


def xls_head(xlsx_workbook, sheet_name, min_date, max_date):
    xlsx_workbook.create_sheet(sheet_name, index=0)
    sheet = xlsx_workbook[sheet_name]

    sheet.merge_cells("B1:E1")
    # font = openpyxl.styles.Font(name='Arial', size=24, italic=True, color='FF0000')
    font = openpyxl.styles.Font(bold=True)
    sheet['B2'].font = font
    sheet['C2'].font = font
    sheet['D2'].font = font
    sheet['E2'].font = font
    cell = sheet.cell(row=1, column=2)
    cell.value = "Date: " + str(datetime.fromtimestamp(min_date)) + " - " + str(datetime.fromtimestamp(max_date))
    cell = sheet.cell(row=2, column=2)
    cell.value = "Link"
    cell = sheet.cell(row=2, column=3)
    cell.value = "Size"
    cell = sheet.cell(row=2, column=4)
    cell.value = "Requests"
    cell = sheet.cell(row=2, column=5)
    cell.value = "Req/h"
    return sheet


def xls_set_column_size(sheet, start_column_letter, *columns):
    """
        Устанавливает размеры колонок на странице. Колонки должны быть в диапозоне от 'A' до 'Z'
    """
    column_letter = start_column_letter
    good_letters = [letter for letter in range(ord('A'), ord('Z'))]

    for column_size in columns:
        if ord(column_letter) not in good_letters:
            break
        sheet.column_dimensions[column_letter].width = column_size
        column_index = ord(column_letter)
        column_index += 1
        column_letter = chr(column_index)
    return


def time_to_timestump(date_time):
    return int(datetime.strptime(date_time, '%d/%b/%Y:%H:%M:%S').strftime("%s"))


def get_network_traffic_size(byte=0):
    if byte > 1024 * 1024:
        byte = byte / 1024 / 1024
        return str(round(byte, 1)) + ' Mbytes'
    if byte > 1024:
        byte = byte / 1024
        return str(round(byte, 1)) + ' Kbytes'
    return str(byte) + ' Bytes'


def split_strings(rows):
    """
        Функиця разбивает строку на элементы списка по пробелам и формирует множество Имён пользователей
        и дат(по дням).
        Пример:
            192.168.0.248 - servers [31/May/2019:00:01:00 +0500] "CONNECT https://195.122.177.135:443/ HTTP/1.0" 200 685
            ['192.168.0.248', '-', 'servers', '[31/May/2019:00:01:00', '+0500]', '"CONNECT', 'https://195.122.177.135:443/', 'HTTP/1.0"', '200', '685', '']
        Возвращает множество Имён пользователей и дат
    """
    users = set()  # список пользователей
    dates = set()  # список дат по дням

    print("Количество строк:", len(rows))
    for i in range(len(rows)):
        rows[i] = rows[i].split(" ")
        users.add(rows[i][USER])

        date = rows[i][DATE][1::]
        rows[i][DATE] = date
        dates.add(date.split(":")[0])

    users.add(USERNAME_ALL_USERS)
    users = sorted(users, key=lambda x: x[0])
    dates = sorted(dates)
    return users, dates


def parse_domain_name(link):
    """
        Получет доменное имя без http://, https://.
        Уберает '/' и ':443' из строки
    """
    if link.startswith("http:") or link.startswith("https:"):
        link = link.split("//")[1]
        link = link.split("/")[0]
    if link.endswith(":443"):
        link = link.split(":443")[0]
    return link


main()
