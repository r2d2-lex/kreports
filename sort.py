#!/usr/bin/env python3
import random
import re
import collections
import requests
from html.parser import HTMLParser
import openpyxl
from datetime import datetime
import time
import locale
import os
import settings

# 192.168.0.248', '-', 'servers', '[25/Mar/2019:01:18:25', '+0500]', '"CONNECT', 'https://195.122.177.135:443/', 'HTTP/1.0"', '200', '856
USERNAME_ALL_USERS = 'All'
REPORT_FILENAME = 'example.xlsx'
WEBTIMEOUT = 60  # время в минутах счётчика посещений
COUNT, SIZEB, UXTIME, VISITS = (0, 1, 2, 3)
IP, USER, DATE, LINK, BYTES = (0, 2, 3, 6, 9)


def main():
    requests.packages.urllib3.disable_warnings()

    filename = choose_log()
    with open(filename, 'r') as fs:
        rows = fs.read().splitlines()

    users, dates = split_strings(rows)
    print("Пользователи:", users)
    print("Даты:", dates)
    try:
        print('Введите имя пользователя или \'{}\' если '
              'хотите выбрать всех пользователей...'.format(USERNAME_ALL_USERS))
        username = input("Введите Имя пользователя: ")
    except KeyboardInterrupt:
        exit()

    print('Текущий пользователь - ', username)
    if username not in users:
        print("Неверное имя пользователя")
        exit()

    generate_report(REPORT_FILENAME, rows, username)
    return


def generate_report(report_filename, rows, username):
    row = 0  # счётчик строк лога
    lst = {}  # словарь пользователя
    min_date = 999999999999999
    max_date = 0

    while row < len(rows):
        try:
            user = rows[row][USER]
            link = rows[row][LINK]
            link = parse_domain_name(link)
            byte = int(rows[row][BYTES])
            time_stump = time_to_timestump(rows[row][DATE])
            min_date = time_stump if time_stump < min_date else min_date
            max_date = time_stump if time_stump > max_date else max_date
            row += 1

            if username != "All":
                if user != username:
                    continue

            db_record = lst.get(user)
            if db_record:
                db_record = lst.pop(user)
                lst[user] = CheckLinkExists(db_record, link, byte, time_stump)
            else:
                # First user record - Count,Size,UnixTime,Visits
                links = {
                    link: [1, byte, time_stump, 1],
                }
                lst[user] = links
        except EOFError:
            break

    wb = openpyxl.Workbook()

    for user in sorted(lst.keys(), reverse=True):
        if user in settings.EXCLUDE:
            continue
        data = lst[user]
        print("___ User: ", user, " Items:,", len(data), "______________________________")
        sheet = xlsHead(wb, user, min_date, max_date)
        lenlink, lensize = 0, 0
        xrow = 3  # Стартовая строка для записи логов

        for lnk, opts in sorted(data.items(), reverse=True, key=lambda x: x[1][1]):
            print("Link: {0}  Options: {1}  Traff:{2}".format(lnk, opts, traf(opts[SIZEB])))
            if opts[SIZEB] == 0:
                continue
            if len(lnk) > lenlink:
                lenlink = len(lnk)
            if len(traf(opts[SIZEB])) > lensize:
                lensize = len(traf(opts[SIZEB]))
            xlsInsert(sheet, xrow, 2, lnk, traf(opts[SIZEB]), opts[COUNT], opts[VISITS])
            xrow += 1
        xlsSetColumn(sheet, lenlink, lensize, 10, 10)
    print("MinDate", datetime.fromtimestamp(min_date), " MaxDate: ", datetime.fromtimestamp(max_date))
    wb.save(report_filename)


def choose_log():
    """
        Ищет файлы срасширением *.log. Если находит, то просит указать индекс лога.
        Если не находит, то просит ввести имя файла.
        На выходе - имя файла для открытия
    """
    while True:
        files = [x for x in os.listdir(".") if x.endswith(".log")]
        if not files:
            filename = input("Введите имя файла с логом: ")
            if not filename:
                print('Введите корректное имя файла')
                continue
            return filename

        dict_of_log_files = {}
        file_index = 0
        for file_name in sorted(files, reverse=True):
            dict_of_log_files[file_index] = file_name
            print("{0}: {1}".format(file_index, file_name))
            file_index += 1

        try:
            index = int(input("Введите индекс лог-файла:"))
            filename = dict_of_log_files[index]
            return filename
        except (KeyError, ValueError):
            print("Введите корректный индекс лог-файла")
            continue


def xlsInsert(sheet, xrow, xcol, link, traf, req, visits):
    """
    sheet - xls sheet,(xrow,xcol) - row + column, link - link of resource, traf - traffic from link,
    req - numbers os requests from browser, vivists - visits in hour
    """
    cell = sheet.cell(row=xrow, column=xcol)
    cell.value = link
    xcol += 1
    cell = sheet.cell(row=xrow, column=xcol)
    cell.value = traf
    xcol += 1
    cell = sheet.cell(row=xrow, column=xcol)
    cell.value = req
    xcol += 1
    cell = sheet.cell(row=xrow, column=xcol)
    cell.value = visits
    # if lnk.startswith("http:"):
    #    print(GetTitle(lnk))
    return


def xlsHead(wb, listname, mindate, maxdate):
    wb.create_sheet(listname, index=0)
    sheet = wb[listname]
    xrow = 3
    sheet.merge_cells("B1:E1")
    # font = openpyxl.styles.Font(name='Arial', size=24, italic=True, color='FF0000')
    font = openpyxl.styles.Font(bold=True)
    sheet['B2'].font = font
    sheet['C2'].font = font
    sheet['D2'].font = font
    sheet['E2'].font = font
    cell = sheet.cell(row=1, column=2)
    cell.value = "Date: " + str(datetime.fromtimestamp(mindate)) + " - " + str(datetime.fromtimestamp(maxdate))
    cell = sheet.cell(row=2, column=2)
    cell.value = "Link"
    cell = sheet.cell(row=2, column=3)
    cell.value = "Size"
    cell = sheet.cell(row=2, column=4)
    cell.value = "Requests"
    cell = sheet.cell(row=2, column=5)
    cell.value = "Req/h"
    return sheet


def xlsSetColumn(sheet, first, second, third, four):
    sheet.column_dimensions['B'].width = first
    sheet.column_dimensions['C'].width = second
    sheet.column_dimensions['D'].width = third
    sheet.column_dimensions['E'].width = four
    return


def time_to_timestump(date_time):
    date_time = date_time[1::]
    return int(datetime.strptime(date_time, '%d/%b/%Y:%H:%M:%S').strftime("%s"))


def traf(byte=0):
    if byte > 1024 * 1024:
        byte = byte / 1024 / 1024
        return (str(round(byte, 1)) + ' Mbytes')
    if byte > 1024:
        byte = byte / 1024
        return (str(round(byte, 1)) + ' Kbytes')
    return (str(byte) + ' Bytes')


def CheckLinkExists(dbres, link, byte, uxtime):
    for dblink, options in dbres.items():
        if dblink == link:
            # Count links
            count = options[COUNT]
            count += 1
            # Size of link
            sizeb = options[SIZEB]
            sizeb += byte
            # Visits
            visits = int(options[VISITS])
            # UxTime
            dbuxtime = int(options[UXTIME])
            if (uxtime - dbuxtime > 60 * WEBTIMEOUT):
                dbuxtime = uxtime
                visits += 1
            dbres[link] = [count, sizeb, dbuxtime, visits]
            return (dbres)
    else:
        dbres[link] = [1, byte, uxtime, 1]
    return (dbres)


def split_strings(rows):
    """
        Функиця разбивает строку на элементы списка по пробелам и формирует множество Имён пользователей
        и дат(по дням).
        Пример:
            192.168.0.248 - servers [31/May/2019:00:01:00 +0500] "CONNECT https://195.122.177.135:443/ HTTP/1.0" 200 685
            ['192.168.0.248', '-', 'servers', '[31/May/2019:00:01:00', '+0500]', '"CONNECT', 'https://195.122.177.135:443/', 'HTTP/1.0"', '200', '685', '']
        Возвращает множество Имён пользователей и дат
    """
    users = set()  # список пользователей
    dates = set()  # список дат по дням

    print("Количество строк:", len(rows))
    for i in range(len(rows)):
        rows[i] = rows[i].split(" ")
        users.add(rows[i][USER])

        date = rows[i][DATE][1::]
        rows[i][DATE] = date
        dates.add(date.split(":")[0])

    users.add(USERNAME_ALL_USERS)
    users = sorted(users, key=lambda x: x[0])
    dates = sorted(dates)
    return users, dates


def parse_domain_name(link):
    """
        Получет доменное имя без http://, https://.
        Уберает '/' и ':443' из строки
    """
    if link.startswith("http:") or link.startswith("https:"):
        link = link.split("//")[1]
        link = link.split("/")[0]
    if link.endswith(":443"):
        link = link.split(":443")[0]
    return link


class MyHTMLParser(HTMLParser):
    def handle_endtag(self, tag):
        if tag == 'title':
            raise StopIteration()

    def handle_data(self, data):
        self.title = data


def GetTitle(url):
    http_proxy = "http://192.168.1.1:8080"
    https_proxy = "https://192.168.1.1:8080"
    ftp_proxy = "ftp://192.168.1.1:8080"
    proxyDict = {
        "http": http_proxy,
        "https": https_proxy,
        "ftp": ftp_proxy
    }
    try:
        r = requests.get(url, stream=True, proxies=proxyDict, verify=False, timeout=1)  # включаем потоковый режим
        data = next(r.iter_content(
            2048))  # запрашиваем ровно 512 байт, для чтения тега head этого должно хватать, или можно еще увеличить
        # print("!!!!",data,"!!!!")
        parser = MyHTMLParser()
        if r.encoding is None:
            r.close()
            return None
        sdata = data.decode(r.encoding)
        r.close()
    except requests.exceptions.SSLError:
        return None
    except StopIteration:
        return None
    except Exception:
        return None

    try:
        parser.feed(sdata)
    except StopIteration:
        print(parser.title)
    except Exception:
        return None
    return


main()
